from __future__ import annotations

from pathlib import Path
import re
import math
from shutil import copyfile

import pandas as pd
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
ORDER_TEMPLATE_PATH = BASE_DIR / "inputs" / "ORDER IMPORT TEMPLATE.xlsx"
ITEM_TEMPLATE_PATH = BASE_DIR / "inputs" / "ITEM IMPORT TEMPLATE.xlsx"
VENDORS_PATH = BASE_DIR / "inputs" / "vendors.csv"
MAPPING_PATH = BASE_DIR / "inputs" / "4d_build_code_to_assembly_id_mapping.csv"
ASSEMBLY_IMPORT_PATH = BASE_DIR / "outputs" / "Assembly_Import.xlsx"
PARTS_SUMMARY_PATH = BASE_DIR / "outputs" / "Parts_Summary.csv"
BUILD_CODE_MAP_PATH = BASE_DIR.parent / "Fuzor_Mapper" / "outputs" / "Revit_4D_Build_Code_Map.csv"
MICRO_SCHEDULE_PATH = BASE_DIR.parent / "Micro_Schedule_Generator" / "outputs" / "Micro_Schedule.csv"
OUTPUTS_DIR = BASE_DIR / "outputs"
OUTPUT_XLSX_PATH = OUTPUTS_DIR / "Production_Order.xlsx"
ITEM_OUTPUT_XLSX_PATH = OUTPUTS_DIR / "Production_Order_Items.xlsx"
ASSEMBLY_PUSH_MAP_PATH = OUTPUTS_DIR / "Revit_Assembly_Id_Map.csv"
KIT_PUSH_MAP_PATH = OUTPUTS_DIR / "Revit_Kit_Parameter_Map.csv"
CENTRAL_BIM_CONTEXT_PATH = BASE_DIR.parent.parent.parent / "outputs" / "takt_zones" / "central_bim_model_llm_context.csv"

CONCRETE_TRUCK_VOLUME_CF = 400
CONCRETE_TRUCK_ASSEMBLY_ID = "CONC_400CF_TRUCK_ASSEMBLY"
FOUNDATION_CONCRETE_TASKS = {
    "Footings - Form/Rebar/Pour",
    "Basement Retaining Walls",
    "Basement Slab on Grade",
    "Foundational Columns",
    "Grade Beams",
    "Rocking Walls",
}

OUTPUT_COLUMNS = [
    "Order ID",
    "Template ID",
    "Description",
    "Order Name",
    "Location",
    "Detailby",
    "Manufactureby",
    "Onsite",
    "Notes",
]

ITEM_OUTPUT_COLUMNS = [
    "Order ID",
    "Item Name",
    "Quantity",
    "Catalog ID",
    "Item ID",
    "Notes",
]


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def slugify_token(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in value.upper())
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    return cleaned.strip("_") or "UNKNOWN"


def level_token(value: object) -> str:
    text = clean_text(value) or "UNASSIGNED"
    text = text.replace("-", " NEG ")
    return slugify_token(text)


def load_template_columns() -> list[str]:
    template = pd.read_excel(ORDER_TEMPLATE_PATH)
    return [str(column).strip() for column in template.columns]


def load_item_template_columns() -> list[str]:
    template = pd.read_excel(ITEM_TEMPLATE_PATH)
    return [str(column).strip() for column in template.columns]


def load_template_id_lookup() -> dict[str, str]:
    vendors = pd.read_csv(VENDORS_PATH, dtype=str).fillna("")
    vendors.columns = [clean_text(column).replace("_", " ") for column in vendors.columns]
    expected = {"Template ID", "Building System"}
    if not expected.issubset(vendors.columns):
        raise ValueError(
            f"Vendor CSV must contain columns {sorted(expected)}; found {list(vendors.columns)}"
        )
    lookup: dict[str, str] = {}
    for _, row in vendors.iterrows():
        template_id = clean_text(row["Template ID"])
        building_system = clean_text(row["Building System"])
        if template_id and building_system:
            lookup[building_system.casefold()] = template_id
    return lookup


def load_mapping() -> pd.DataFrame:
    mapping = pd.read_csv(MAPPING_PATH, dtype=str).fillna("")
    expected = {"build_code", "assembly_id"}
    if not expected.issubset(mapping.columns):
        raise ValueError(
            f"Mapping CSV must contain columns {sorted(expected)}; found {list(mapping.columns)}"
        )
    mapping["build_code"] = mapping["build_code"].apply(clean_text)
    mapping["assembly_id"] = mapping["assembly_id"].apply(clean_text)
    mapping["element_id"] = ""
    mapping = mapping[(mapping["build_code"] != "") & (mapping["assembly_id"] != "")].copy()
    structural_mapping = load_structural_mapping()
    dynamic_mapping = load_dynamic_mapping()
    return pd.concat(
        [mapping.loc[:, ["build_code", "assembly_id", "element_id"]], structural_mapping, dynamic_mapping],
        ignore_index=True,
    )


def load_dynamic_mapping() -> pd.DataFrame:
    if not MICRO_SCHEDULE_PATH.exists():
        return pd.DataFrame(columns=["build_code", "assembly_id", "element_id"])

    rows: list[dict[str, str]] = []
    existing_mapped_build_codes: set[str] = set()
    if MAPPING_PATH.exists():
        static_mapping = pd.read_csv(MAPPING_PATH, dtype=str).fillna("")
        if {"build_code", "assembly_id"}.issubset(static_mapping.columns):
            existing_mapped_build_codes = {
                clean_text(row["build_code"])
                for _, row in static_mapping.iterrows()
                if clean_text(row["build_code"]) and clean_text(row["assembly_id"])
            }

    if BUILD_CODE_MAP_PATH.exists():
        build_df = pd.read_csv(BUILD_CODE_MAP_PATH, dtype=str).fillna("")
        exterior = build_df[build_df["task_name"].astype(str).str.strip().eq("Exterior Wall Install")].copy()
        for build_code, group in exterior.groupby("build_code", sort=False):
            build_code = clean_text(build_code)
            if not build_code or build_code in existing_mapped_build_codes:
                continue
            build_token = slugify_token(build_code.split("|")[-1])
            rows.append(
                {
                    "build_code": build_code,
                    "assembly_id": f"EXTW_{build_token}",
                    "element_id": "",
                }
            )

        interior = build_df[build_df["task_name"].astype(str).str.strip().eq("Interior Walls")].copy()
        for _, row in interior.drop_duplicates(subset=["element_id", "build_code"]).iterrows():
            build_code = clean_text(row.get("build_code"))
            level = clean_text(row.get("level")) or "Unassigned"
            element_id = clean_text(row.get("element_id"))
            if not build_code or not element_id:
                continue
            rows.append(
                {
                    "build_code": build_code,
                    "assembly_id": f"INTW_{level_token(level)}_ASSEMBLY",
                    "element_id": element_id,
                }
            )

        mesh = build_df[build_df["task_name"].astype(str).str.strip().eq("Mesh Install")].copy()
        for _, row in mesh.drop_duplicates(subset=["element_id", "build_code"]).iterrows():
            build_code = clean_text(row.get("build_code"))
            level = clean_text(row.get("level")) or "Unassigned"
            element_id = clean_text(row.get("element_id"))
            if not build_code or not element_id:
                continue
            rows.append(
                {
                    "build_code": build_code,
                    "assembly_id": f"MESH_{level_token(level)}_ASSEMBLY",
                    "element_id": element_id,
                }
            )

    for group in build_mep_groups():
        rows.append(
            {
                "build_code": group["build_code"],
                "assembly_id": group["assembly_id"],
                "element_id": "",
            }
        )

    for group in build_concrete_truck_groups():
        rows.append(
            {
                "build_code": group["build_code"],
                "assembly_id": CONCRETE_TRUCK_ASSEMBLY_ID,
                "element_id": "",
            }
        )

    if not rows:
        return pd.DataFrame(columns=["build_code", "assembly_id", "element_id"])
    return pd.DataFrame(rows).drop_duplicates().reset_index(drop=True)


def build_mep_groups() -> list[dict[str, object]]:
    if not MICRO_SCHEDULE_PATH.exists():
        return []

    micro = pd.read_csv(MICRO_SCHEDULE_PATH, dtype=str).fillna("")
    mep = micro[micro["task_name"].astype(str).str.strip().eq("MEP Rough-In")].copy()
    if mep.empty:
        return []

    mep["element_start_ts"] = pd.to_datetime(mep["element_start"], format="mixed", errors="coerce")
    mep_unique = (
        mep.sort_values(["level", "element_start_ts", "order_in_task", "element_id"])
        .drop_duplicates(subset=["element_id"])
        .reset_index(drop=True)
    )

    groups: list[dict[str, object]] = []
    for level, level_df in mep_unique.groupby("level", sort=False):
        level_text = clean_text(level) or "Unassigned"
        token = level_token(level_text)
        level_df = level_df.reset_index(drop=True)
        for group_index, start in enumerate(range(0, len(level_df), 3), start=1):
            group = level_df.iloc[start : start + 3].copy()
            build_code = f"MEP Rough-In | {level_text} | MEP_{token}_G{group_index:03d}"
            groups.append(
                {
                    "build_code": build_code,
                    "assembly_id": f"MEP_{token}_G{group_index:03d}",
                    "elements": group,
                }
            )
    return groups


def parse_quantity(value: object) -> float:
    text = clean_text(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return max(float(match.group()), 0.0) if match else 0.0


def load_element_volume_lookup() -> dict[str, float]:
    if not CENTRAL_BIM_CONTEXT_PATH.exists():
        return {}
    context = pd.read_csv(CENTRAL_BIM_CONTEXT_PATH, dtype=str).fillna("")
    if not {"element_id", "volume"}.issubset(context.columns):
        return {}
    return {
        clean_text(row["element_id"]): parse_quantity(row["volume"])
        for _, row in context.iterrows()
        if clean_text(row["element_id"])
    }


def build_concrete_truck_groups() -> list[dict[str, object]]:
    if not BUILD_CODE_MAP_PATH.exists():
        return []

    build_df = pd.read_csv(BUILD_CODE_MAP_PATH, dtype=str).fillna("")
    concrete = build_df[build_df["task_name"].isin(FOUNDATION_CONCRETE_TASKS)].copy()
    if concrete.empty:
        return []

    volume_lookup = load_element_volume_lookup()
    groups: list[dict[str, object]] = []
    for (task_name, level), group in concrete.groupby(["task_name", "level"], sort=False):
        group = group.drop_duplicates(subset=["element_id"]).copy()
        group["volume_cf"] = group["element_id"].map(lambda value: volume_lookup.get(clean_text(value), 0.0))
        total_volume = float(group["volume_cf"].sum())
        truck_count = max(math.ceil(total_volume / CONCRETE_TRUCK_VOLUME_CF), 1)
        task_token = slugify_token(clean_text(task_name))
        level_text = clean_text(level) or "Unassigned"
        level_slug = level_token(level_text)
        for truck_index in range(1, truck_count + 1):
            truck_build_code = f"{task_name} Concrete | {level_text} | {task_token}_{level_slug}_TRUCK_{truck_index:03d}"
            truck_group = group.copy()
            truck_group["element_start"] = pd.to_datetime(truck_group["element_start"], format="mixed", errors="coerce")
            groups.append(
                {
                    "build_code": truck_build_code,
                    "assembly_id": CONCRETE_TRUCK_ASSEMBLY_ID,
                    "elements": truck_group,
                    "total_volume_cf": total_volume,
                    "truck_count": truck_count,
                }
            )
    return groups


def load_structural_mapping() -> pd.DataFrame:
    if not PARTS_SUMMARY_PATH.exists() or not BUILD_CODE_MAP_PATH.exists():
        return pd.DataFrame(columns=["build_code", "assembly_id", "element_id"])

    summary = pd.read_csv(PARTS_SUMMARY_PATH, dtype=str).fillna("")
    build_df = pd.read_csv(BUILD_CODE_MAP_PATH, dtype=str).fillna("")
    if not {"element_id", "build_code"}.issubset(build_df.columns):
        return pd.DataFrame(columns=["build_code", "assembly_id", "element_id"])

    build_lookup = {
        clean_text(row["element_id"]): clean_text(row["build_code"])
        for _, row in build_df.iterrows()
        if clean_text(row["element_id"]) and clean_text(row["build_code"])
    }

    rows: list[dict[str, str]] = []
    structural_summary = summary[summary["part_id"].astype(str).str.startswith("STRUCT_")].copy()
    for _, part_row in structural_summary.iterrows():
        assembly_id = clean_text(part_row["part_id"])
        for element_id in clean_text(part_row.get("element_ids", "")).split("|"):
            element_id = clean_text(element_id)
            build_code = build_lookup.get(element_id, "")
            if not assembly_id or not element_id or not build_code:
                continue
            rows.append(
                {
                    "build_code": build_code,
                    "assembly_id": assembly_id,
                    "element_id": element_id,
                }
            )

    if not rows:
        return pd.DataFrame(columns=["build_code", "assembly_id", "element_id"])

    return (
        pd.DataFrame(rows)
        .drop_duplicates(subset=["build_code", "assembly_id", "element_id"])
        .sort_values(["assembly_id", "element_id"])
        .reset_index(drop=True)
    )


def load_assembly_lookup() -> dict[str, dict[str, str]]:
    assembly_df = pd.read_excel(ASSEMBLY_IMPORT_PATH).fillna("")
    assembly_df["ID"] = assembly_df["ID"].apply(clean_text)
    assembly_df = assembly_df[assembly_df["ID"] != ""].copy()
    lookup: dict[str, dict[str, str]] = {}
    for _, row in assembly_df.iterrows():
        lookup[clean_text(row["ID"])] = {
            "name": clean_text(row["Name"]),
            "catalog_id": clean_text(row["Catalog Id"]),
            "description": clean_text(row["Description"]),
            "category": clean_text(row.get("Category", "")),
            "sub_category": clean_text(row.get("Sub Category", "")),
        }
    return lookup


def load_build_code_starts() -> dict[str, pd.Timestamp]:
    build_df = load_build_code_elements()
    build_df["element_start"] = pd.to_datetime(build_df["element_start"], format="mixed", errors="coerce")
    starts = (
        build_df.groupby("build_code", as_index=False)["element_start"]
        .min()
        .set_index("build_code")["element_start"]
        .to_dict()
    )
    return starts


def load_build_code_elements() -> pd.DataFrame:
    build_df = pd.read_csv(BUILD_CODE_MAP_PATH, dtype=str).fillna("")
    required = {"element_id", "build_code"}
    if not required.issubset(build_df.columns):
        raise ValueError(
            f"4D build-code map is missing required columns: {sorted(required - set(build_df.columns))}"
        )
    build_df["element_id"] = build_df["element_id"].apply(clean_text)
    build_df["build_code"] = build_df["build_code"].apply(clean_text)
    if "element_start" in build_df.columns:
        build_df["element_start"] = pd.to_datetime(build_df["element_start"], format="mixed", errors="coerce")
    build_df = build_df[(build_df["element_id"] != "") & (build_df["build_code"] != "")].copy()

    dynamic_rows: list[dict[str, object]] = []
    for group in build_mep_groups():
        group_build_code = clean_text(group["build_code"])
        group_elements = group["elements"]
        if not isinstance(group_elements, pd.DataFrame):
            continue
        for _, row in group_elements.iterrows():
            dynamic_rows.append(
                {
                    "element_id": clean_text(row.get("element_id")),
                    "build_code": group_build_code,
                    "task_id": clean_text(row.get("task_id")),
                    "task_name": clean_text(row.get("task_name")),
                    "level": clean_text(row.get("level")),
                    "element_key": clean_text(row.get("element_key")),
                    "element_start": pd.to_datetime(row.get("element_start"), errors="coerce"),
                    "element_end": clean_text(row.get("element_end")),
                }
            )

    if dynamic_rows:
        build_df = pd.concat([build_df, pd.DataFrame(dynamic_rows)], ignore_index=True)

    concrete_rows: list[dict[str, object]] = []
    for group in build_concrete_truck_groups():
        group_build_code = clean_text(group["build_code"])
        group_elements = group["elements"]
        if not isinstance(group_elements, pd.DataFrame):
            continue
        for _, row in group_elements.iterrows():
            concrete_rows.append(
                {
                    "element_id": clean_text(row.get("element_id")),
                    "build_code": group_build_code,
                    "task_id": clean_text(row.get("task_id")),
                    "task_name": f"{clean_text(row.get('task_name'))} Concrete",
                    "level": clean_text(row.get("level")),
                    "element_key": clean_text(row.get("element_key")),
                    "element_start": pd.to_datetime(row.get("element_start"), errors="coerce"),
                    "element_end": clean_text(row.get("element_end")),
                }
            )

    if concrete_rows:
        build_df = pd.concat([build_df, pd.DataFrame(concrete_rows)], ignore_index=True)

    return build_df[(build_df["element_id"] != "") & (build_df["build_code"] != "")].copy()


def kit_id_from_assembly_and_onsite(assembly_id: str, onsite: pd.Timestamp) -> str:
    return f"{assembly_id}_{onsite.strftime('%Y%m%d')}"


def natural_order_key(assembly_id: str, onsite: pd.Timestamp) -> str:
    return f"{assembly_id}|{onsite.strftime('%Y%m%d')}"


def location_from_build_code(build_code: str) -> str:
    parts = [clean_text(part) for part in build_code.split("|")]
    parts = [part for part in parts if part]
    if len(parts) >= 2:
        return parts[1]
    return build_code


def building_system_for_assembly(assembly_id: str, assembly: dict[str, str]) -> str:
    if assembly_id.startswith("STRUCT_"):
        return "Superstructure"
    return assembly["sub_category"] or assembly["category"]


def template_details(assembly_id: str, assembly: dict[str, str], template_id_lookup: dict[str, str]) -> tuple[str, str]:
    building_system = building_system_for_assembly(assembly_id, assembly)
    template_id = template_id_lookup.get(building_system.casefold(), "")
    if not template_id and len(template_id_lookup) == 1:
        template_id = next(iter(template_id_lookup.values()))
    if not template_id:
        raise ValueError(f"No Template ID found in vendors.csv for building system `{building_system}`")
    return template_id, building_system


def short_description(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9 ]+", " ", clean_text(value))
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        text = "Order"
    text = text[:24].strip()
    text = re.sub(r"^[^A-Za-z0-9]+|[^A-Za-z0-9]+$", "", text)
    if len(text) < 4:
        text = f"{text} Order".strip()
    return text[:24].strip()


def build_kit_import(
    mapping: pd.DataFrame,
    assembly_lookup: dict[str, dict[str, str]],
    build_code_starts: dict[str, pd.Timestamp],
    template_id_lookup: dict[str, str],
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []

    for _, row in mapping.iterrows():
        build_code = clean_text(row["build_code"])
        assembly_id = clean_text(row["assembly_id"])
        if build_code not in build_code_starts:
            raise ValueError(f"Build code not found in 4D build-code map: `{build_code}`")
        if assembly_id not in assembly_lookup:
            raise ValueError(f"Assembly ID not found in Assembly_Import.xlsx: `{assembly_id}`")

        onsite = pd.Timestamp(build_code_starts[build_code]).normalize()
        manufacture = onsite - pd.Timedelta(days=2)
        detail = onsite - pd.Timedelta(days=4)
        assembly = assembly_lookup[assembly_id]
        template_id, building_system = template_details(assembly_id, assembly, template_id_lookup)
        building_system_note = f" | Building System: {building_system}" if building_system else ""
        description = short_description(assembly["name"] or assembly_id)
        order_key = natural_order_key(assembly_id, onsite)

        rows.append(
            {
                "_Natural Order Key": order_key,
                "_Assembly ID": assembly_id,
                "_Assembly Catalog ID": assembly["catalog_id"],
                "_Assembly Name": assembly["name"] or assembly_id,
                "_Build Code": build_code,
                "Order ID": "",
                "Template ID": template_id,
                "Description": description,
                "Order Name": assembly["name"] or assembly_id,
                "Location": location_from_build_code(build_code),
                "Detailby": detail.date(),
                "Manufactureby": manufacture.date(),
                "Onsite": onsite.date(),
                "Notes": f"Assembly ID: {assembly_id} | Assembly Catalog Id: {assembly['catalog_id']} | 4D Build Code: {build_code}{building_system_note}",
            }
        )

    output = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
    output = pd.DataFrame(rows)
    if output.empty:
        return output

    def combine_notes(series: pd.Series) -> str:
        values = [clean_text(value) for value in series if clean_text(value)]
        return " || ".join(dict.fromkeys(values))

    aggregated = (
        output.groupby("_Natural Order Key", as_index=False)
        .agg(
            {
                "_Assembly ID": "first",
                "_Assembly Catalog ID": "first",
                "_Assembly Name": "first",
                "_Build Code": "nunique",
                "Template ID": "first",
                "Description": "first",
                "Order Name": "first",
                "Location": "first",
                "Detailby": "first",
                "Manufactureby": "first",
                "Onsite": "first",
                "Notes": combine_notes,
            }
        )
        .sort_values(["Onsite", "_Assembly ID"])
        .reset_index(drop=True)
    )
    aggregated["Order ID"] = [f"PO-{index:06d}" for index in range(1, len(aggregated) + 1)]
    return aggregated


def build_item_import(order_output: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for _, row in order_output.iterrows():
        quantity = int(row["_Build Code"])
        rows.append(
            {
                "Order ID": row["Order ID"],
                "Item Name": row["_Assembly Name"],
                "Quantity": quantity,
                "Catalog ID": row["_Assembly Catalog ID"],
                "Item ID": row["_Assembly ID"],
                "Notes": row["Notes"],
            }
        )
    return pd.DataFrame(rows, columns=ITEM_OUTPUT_COLUMNS)


def build_push_maps(
    mapping: pd.DataFrame,
    kit_output: pd.DataFrame,
    build_code_elements: pd.DataFrame,
    assembly_lookup: dict[str, dict[str, str]],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    order_lookup = kit_output.set_index("_Natural Order Key").to_dict(orient="index")
    rows: list[dict[str, object]] = []

    for _, row in mapping.iterrows():
        build_code = clean_text(row["build_code"])
        assembly_id = clean_text(row["assembly_id"])
        element_id_filter = clean_text(row.get("element_id", ""))
        build_rows = build_code_elements[build_code_elements["build_code"] == build_code]
        if element_id_filter:
            build_rows = build_rows[build_rows["element_id"] == element_id_filter]
        if build_rows.empty:
            continue

        onsite = pd.Timestamp(build_rows.get("element_start", pd.Series(dtype=str)).min()).normalize() if "element_start" in build_rows.columns else None
        if onsite is None or pd.isna(onsite):
            raise ValueError(f"Could not determine onsite date for build code `{build_code}`")
        order_key = natural_order_key(assembly_id, onsite)
        order_data = order_lookup.get(order_key)
        if not order_data:
            raise ValueError(f"Order data not found for `{order_key}`")
        order_id = clean_text(order_data["Order ID"])
        item_name = clean_text(order_data["_Assembly Name"])

        for _, build_row in build_rows.iterrows():
            element_id = clean_text(build_row["element_id"])
            if not element_id:
                continue
            rows.append(
                {
                    "element_id": element_id,
                    "assembly_id": assembly_id,
                    "catalog_id": clean_text(assembly_lookup[assembly_id]["catalog_id"]),
                    "kit_id": kit_id_from_assembly_and_onsite(assembly_id, onsite),
                    "order_id": order_id,
                    "item_name": item_name,
                    "build_code": build_code,
                }
            )

    push_df = pd.DataFrame(rows)
    if push_df.empty:
        return (
            pd.DataFrame(columns=["element_id", "assembly_id", "catalog_id", "build_code"]),
            pd.DataFrame(columns=["element_id", "kit_id", "order_id", "item_name", "build_code"]),
        )

    assembly_push = (
        push_df.loc[:, ["element_id", "assembly_id", "catalog_id", "build_code"]]
        .drop_duplicates(subset=["element_id", "assembly_id", "catalog_id", "build_code"])
        .sort_values(["assembly_id", "element_id"])
        .reset_index(drop=True)
    )
    kit_push = (
        push_df.loc[:, ["element_id", "kit_id", "order_id", "item_name", "build_code"]]
        .drop_duplicates(subset=["element_id", "kit_id", "order_id", "item_name", "build_code"])
        .sort_values(["kit_id", "element_id"])
        .reset_index(drop=True)
    )
    return assembly_push, kit_push


def write_order_import(output: pd.DataFrame) -> None:
    copyfile(ORDER_TEMPLATE_PATH, OUTPUT_XLSX_PATH)
    workbook = load_workbook(OUTPUT_XLSX_PATH)
    if "ORDERS" not in workbook.sheetnames:
        raise ValueError(f"Order import template must contain an `ORDERS` sheet; found {workbook.sheetnames}")

    sheet = workbook["ORDERS"]
    headers = [clean_text(sheet.cell(row=1, column=column).value) for column in range(1, sheet.max_column + 1)]
    if headers != OUTPUT_COLUMNS:
        raise ValueError(
            "Template columns do not match the expected Order Import structure.\n"
            f"Expected: {OUTPUT_COLUMNS}\n"
            f"Found: {headers}"
        )

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    for row_index, (_, output_row) in enumerate(output.iterrows(), start=2):
        for column_index, column_name in enumerate(OUTPUT_COLUMNS, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=output_row[column_name])
            if column_name in {"Order ID", "Template ID"}:
                cell.number_format = "@"

    workbook.save(OUTPUT_XLSX_PATH)


def write_item_import(output: pd.DataFrame) -> None:
    copyfile(ITEM_TEMPLATE_PATH, ITEM_OUTPUT_XLSX_PATH)
    workbook = load_workbook(ITEM_OUTPUT_XLSX_PATH)
    if "ITEMS" not in workbook.sheetnames:
        raise ValueError(f"Item import template must contain an `ITEMS` sheet; found {workbook.sheetnames}")

    sheet = workbook["ITEMS"]
    headers = [clean_text(sheet.cell(row=1, column=column).value) for column in range(1, sheet.max_column + 1)]
    if headers != ITEM_OUTPUT_COLUMNS:
        raise ValueError(
            "Template columns do not match the expected Item Import structure.\n"
            f"Expected: {ITEM_OUTPUT_COLUMNS}\n"
            f"Found: {headers}"
        )

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    for row_index, (_, output_row) in enumerate(output.iterrows(), start=2):
        for column_index, column_name in enumerate(ITEM_OUTPUT_COLUMNS, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=output_row[column_name])
            if column_name in {"Order ID", "Catalog ID", "Item ID"}:
                cell.number_format = "@"
            elif column_name == "Quantity":
                cell.number_format = "0"

    workbook.save(ITEM_OUTPUT_XLSX_PATH)


if __name__ == "__main__":
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

    template_columns = load_template_columns()
    if template_columns != OUTPUT_COLUMNS:
        raise ValueError(
            "Template columns do not match the expected Order Import structure.\n"
            f"Expected: {OUTPUT_COLUMNS}\n"
            f"Found: {template_columns}"
        )
    item_template_columns = load_item_template_columns()
    if item_template_columns != ITEM_OUTPUT_COLUMNS:
        raise ValueError(
            "Template columns do not match the expected Item Import structure.\n"
            f"Expected: {ITEM_OUTPUT_COLUMNS}\n"
            f"Found: {item_template_columns}"
        )

    mapping = load_mapping()
    assembly_lookup = load_assembly_lookup()
    build_code_starts = load_build_code_starts()
    build_code_elements = load_build_code_elements()
    template_id_lookup = load_template_id_lookup()
    output = build_kit_import(mapping, assembly_lookup, build_code_starts, template_id_lookup)
    item_output = build_item_import(output)
    assembly_push, kit_push = build_push_maps(mapping, output, build_code_elements, assembly_lookup)

    assembly_push.to_csv(ASSEMBLY_PUSH_MAP_PATH, index=False)
    kit_push.to_csv(KIT_PUSH_MAP_PATH, index=False)
    write_item_import(item_output)
    write_order_import(output)

    print(f"Wrote {ITEM_OUTPUT_XLSX_PATH.name}")
    print(f"Wrote {OUTPUT_XLSX_PATH.name}")
    print(f"Wrote {ASSEMBLY_PUSH_MAP_PATH.name}")
    print(f"Wrote {KIT_PUSH_MAP_PATH.name}")
    print(f"Rows: {len(output)}")
    print(f"Item rows: {len(item_output)}")
