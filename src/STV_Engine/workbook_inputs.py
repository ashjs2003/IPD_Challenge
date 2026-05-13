from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


STV_WORKBOOK_CONSTRUCTION_ITEMS = {
    ("Energy", "Photovoltaics (sf)"),
    ("Energy", "EV Battery (kWh)"),
    ("Energy", "Integrated Solar Water Heating (sf)"),
    ("MEP", "Rainwater Collection Tank (gal)"),
}

USE_PHASE_LABELS = {
    "Electricity Drawn from Grid:": ("electricity_from_grid_kwh",),
    "On-site Renewable Electricity:": ("onsite_renewable_kwh",),
    "Natural Gas Use:": ("natural_gas_m3",),
    "Toilet Flow Rate:": ("water_use", "toilet_gpf"),
    "Urinal Flow Rate:": ("water_use", "urinal_gpf"),
    "WC Sink Flow Rate:": ("water_use", "wc_sink_gpm"),
    "Lab Sink Flow Rate:": ("water_use", "lab_sink_gpm"),
    "Kitchen Sink Flow Rate:": ("water_use", "kitchen_sink_gpm"),
    "Shower Flow Rate:": ("water_use", "shower_gpm"),
    "Landscaping Water Use:": ("water_use", "landscaping_gal"),
    "Rainwater Collection:": ("water_use", "rainwater_collection_gal"),
    "Fuel Type": ("cogeneration", "fuel_type"),
    "Electricity": ("cogeneration", "electricity_kwh"),
    "Electricity Split": ("cogeneration", "electricity_split"),
    "Heating": ("cogeneration", "heating_mj"),
    "Heating Split": ("cogeneration", "heating_split"),
    "Cooling": ("cogeneration", "cooling_kwh"),
    "Cooling Split": ("cogeneration", "cooling_split"),
}


def load_stv_workbook_inputs(workbook_path: Path | str) -> dict[str, Any]:
    """Load the limited STV input surface requested from an STV workbook."""
    path = Path(workbook_path)
    wb = load_workbook(path, data_only=True)

    payload: dict[str, Any] = {
        "construction_items": _load_selected_construction_items(wb["Construction and Materials"]),
        "use_phase": _load_use_phase(wb["Use Phase"]),
    }

    team = wb["Construction and Materials"]["C5"].value or wb["Use Phase"]["C5"].value
    if team:
        payload["team"] = str(team)

    return payload


def _load_selected_construction_items(ws) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for row in range(18, ws.max_row + 1):
        assembly = _clean_text(ws.cell(row=row, column=2).value)
        material_type = _clean_text(ws.cell(row=row, column=3).value)
        if (assembly, material_type) not in STV_WORKBOOK_CONSTRUCTION_ITEMS:
            continue

        amount = _to_float(ws.cell(row=row, column=4).value)
        if amount == 0:
            continue

        items.append(
            {
                "assembly": assembly,
                "material_type": material_type,
                "amount": amount,
            }
        )
    return items


def _load_use_phase(ws) -> dict[str, Any]:
    use_phase: dict[str, Any] = {
        "electricity_from_grid_kwh": 0.0,
        "onsite_renewable_kwh": 0.0,
        "natural_gas_m3": 0.0,
        "cogeneration": {
            "fuel_type": None,
            "electricity_kwh": 0.0,
            "heating_mj": 0.0,
            "cooling_kwh": 0.0,
            "electricity_split": 0.0,
            "heating_split": 0.0,
            "cooling_split": 0.0,
        },
        "water_use": {
            "toilet_gpf": 0.0,
            "urinal_gpf": 0.0,
            "wc_sink_gpm": 0.0,
            "lab_sink_gpm": 0.0,
            "kitchen_sink_gpm": 0.0,
            "shower_gpm": 0.0,
            "landscaping_gal": 0.0,
            "rainwater_collection_gal": 0.0,
        },
    }

    for row in range(18, ws.max_row + 1):
        label = _clean_text(ws.cell(row=row, column=2).value) or _clean_text(
            ws.cell(row=row, column=3).value
        )
        path = USE_PHASE_LABELS.get(label)
        if not path:
            continue

        value = ws.cell(row=row, column=4).value
        if path[-1] == "fuel_type":
            fuel_type = _clean_text(value)
            if fuel_type and fuel_type.lower() != "select fuel":
                use_phase["cogeneration"]["fuel_type"] = fuel_type
            continue

        target = use_phase
        for key in path[:-1]:
            target = target[key]
        target[path[-1]] = _to_float(value)

    return use_phase


def _clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _to_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)
